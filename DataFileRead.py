import time

from openpyxl import Workbook, load_workbook

from WebDriver_CrossBrowser import driver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

wb = load_workbook(filename='DataFile.xlsx')
sheet = wb['Sheet']

driver.get("https://www.orangehrm.com/orangehrm-30-day-trial/")
driver.maximize_window()
print(driver.title)
driver.implicitly_wait(5)

"""XPATH"""
userName = driver.find_element(By.XPATH, "//input[@id='Form_submitForm_subdomain']")
name = driver.find_element(By.XPATH, "//input[@id='Form_submitForm_Name']")
email = driver.find_element(By.XPATH, "//input[@id='Form_submitForm_Email']")
phoneNumber = driver.find_element(By.XPATH, "//input[@id='Form_submitForm_Contact']")
country = driver.find_element(By.XPATH, "//select[@id='Form_submitForm_Country']")

rowCount = sheet.max_row
colCount = sheet.max_column

print(rowCount)
print(colCount)

for i in range(1, rowCount + 1):
    userNameValue = sheet.cell(row=i, column=1).value
    nameValue = sheet.cell(row=i, column=2).value
    emailValue = sheet.cell(row=i, column=3).value
    phoneNumberValue = sheet.cell(row=i, column=4).value
    countryValue = sheet.cell(row=i, column=5).value

print(userNameValue, nameValue, emailValue, phoneNumberValue, countryValue)

"""ACTIONS"""
userName.send_keys(userNameValue)
name.send_keys(nameValue)
email.send_keys(emailValue)
phoneNumber.send_keys(phoneNumberValue)

"""DROP_DOWN Method"""


def select_values(element, value):
    select = Select(element)
    select.select_by_visible_text(value)


select_values(country, "Bangladesh")
time.sleep(10)
driver.close()
